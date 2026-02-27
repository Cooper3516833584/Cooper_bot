
# handinsvc.py
from __future__ import annotations
import asyncio
import json
import re
import time
import urllib.request
import urllib.error
import urllib.parse
import shutil
from dataclasses import dataclass, asdict, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Set
from urllib.parse import urlparse, unquote
import os

from config import (
    DATA_DIR,
    GROUP_DOCS_DIR,
    USER_DOCS_DIR,
    HANDIN_DB_PATH,
    HANDIN_INBOX_DIR,
    HANDIN_TASKS_DIRNAME,
    HANDIN_ROOT_DIR,
    ROSTER_XLSX_PATH,
    TIMEZONE,
    NAPCAT_TEMP_CONTAINER_DIR,
    NAPCAT_TEMP_HOST_DIR,
    HANDIN_KEEP_DAYS_AFTER_LAST_GET,
    HANDIN_INBOX_KEEP_DAYS,
)
from logger import Logger

try:
    from zoneinfo import ZoneInfo
except Exception:  # Py<3.9
    ZoneInfo = None

import openpyxl


# ========= æ–‡ä»¶åæå–ï¼ˆå‚è€ƒ who_has_handed_in.py çš„é€»è¾‘ï¼‰ =========
BLACKLIST_SUBSTRINGS = {
    "ç”µæ°”", "å­¦é™¢", "å·¥ç¨‹", "ç­", "ä¸“ä¸š",
    "æŠ¥å‘Š", "è¯»ä¹¦", "ä½œä¸š", "è®ºæ–‡", "é©¬åŸ",
    "è¯¾", "é˜…è¯»", "å†å²", "è‡ªç”±", "ä¹‹é—´",
    "æ”¿æ²»", "ç»æµ", "åºè¨€", "å¯¼è¨€", "ç»å…¸", "æ€æƒ³",
}
STRUCTURAL_WORDS = ["ç”µæ°”", "å­¦é™¢", "å·¥ç¨‹", "ç­", "ä¸“ä¸š"]
SEPARATORS = ["-", "_", "â€”â€”", "â€”", "â€“", ";", "ï¼Œ", ",", " "]

_RE_STU = re.compile(r"[Uu]\d{8,12}")  # ä¾‹å¦‚ U202412743
_RE_ENG = re.compile(r"[A-Za-z]")
_RE_NUM = re.compile(r"[Uu]?\d{4,}")
SUBMITTED_FILE_SUFFIXES = {".doc", ".docx", ".pdf", ".txt", ".zip", ".rar", ".7z", ".ppt", ".pptx", ".xls", ".xlsx"}

def clean_filename(filename: str) -> str:
    stem = Path(filename).stem
    for sep in SEPARATORS:
        stem = stem.replace(sep, " ")
    stem = _RE_NUM.sub(" ", stem)
    stem = _RE_ENG.sub(" ", stem)
    return stem

def looks_like_name(token: str) -> bool:
    if not re.fullmatch(r"[\u4e00-\u9fff]{2,3}", token):
        return False
    for bad in BLACKLIST_SUBSTRINGS:
        if bad in token:
            return False
    return True

def extract_name_from_filename(filename: str) -> str:
    part = clean_filename(filename)
    tokens = [t for t in part.split() if t]
    for tok in reversed(tokens):
        if looks_like_name(tok):
            return tok

    for tok in tokens:
        if not re.fullmatch(r"[\u4e00-\u9fff]+", tok):
            continue
        for sw in STRUCTURAL_WORDS:
            idx = tok.find(sw)
            if idx >= 2:
                prefix = tok[:idx]
                if looks_like_name(prefix):
                    return prefix

    chunks = re.findall(r"[\u4e00-\u9fff]+", part)
    candidates = []
    for chunk in chunks:
        for n in (3, 2):
            for i in range(len(chunk) - n + 1):
                sub = chunk[i:i+n]
                if looks_like_name(sub):
                    candidates.append(sub)
    return candidates[-1] if candidates else ""

def extract_student_id(filename: str) -> str:
    m = _RE_STU.search(filename or "")
    return m.group(0).upper() if m else ""


# ========= åå†Œè¯»å– =========
def load_roster(path: Path) -> List[Tuple[str, str]]:
    """è¯»å–ç­çº§åå†Œï¼Œè¿”å› [(å­¦å·, å§“å), ...]ã€‚æ”¯æŒé¦–è¡Œæ˜¯æ ‡é¢˜ã€ç¬¬äºŒè¡Œæ‰æ˜¯è¡¨å¤´ã€‚"""
    path = Path(path)
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    header_row = None
    col_id = None
    col_name = None

    # æ‰¾åŒ…å«â€œå­¦å·â€â€œå§“åâ€çš„è¡¨å¤´è¡Œ
    for r in range(1, min(30, ws.max_row) + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, min(50, ws.max_column) + 1)]
        # ç»Ÿä¸€æˆå­—ç¬¦ä¸²æ¯”è¾ƒ
        row_str = [str(v).strip() for v in row_vals if v is not None]
        if "å­¦å·" in row_str and "å§“å" in row_str:
            header_row = r
            # æ‰¾åˆ—å·
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                s = str(v).strip()
                if s == "å­¦å·":
                    col_id = c
                elif s == "å§“å":
                    col_name = c
            break

    if not header_row or not col_id or not col_name:
        return []

    out: List[Tuple[str, str]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        sid = ws.cell(r, col_id).value
        name = ws.cell(r, col_name).value
        if sid is None and name is None:
            continue
        sid_s = str(sid).strip() if sid is not None else ""
        name_s = str(name).strip() if name is not None else ""
        if not sid_s and not name_s:
            continue
        out.append((sid_s.upper(), name_s))
    return out


# ========= æ—¶é—´è§£æ =========
_RE_TIME = re.compile(r"^\s*(\d{1,2})[.\u3002/\-](\d{1,2})\s*(\d{1,2})[:ï¼š](\d{1,2})\s*$")

def parse_mmdd_hhmm(s: str, now_ts: float) -> Optional[float]:
    """æŠŠ 'M.D HH:MM' è§£æä¸ºæ—¶é—´æˆ³ã€‚è‹¥è§£æå‡ºçš„æ—¶é—´ <= å½“å‰ï¼Œåˆ™è‡ªåŠ¨ +1 å¹´ã€‚"""
    s = (s or "").strip()
    m = _RE_TIME.match(s)
    if not m:
        return None
    mon = int(m.group(1))
    day = int(m.group(2))
    hh = int(m.group(3))
    mm = int(m.group(4))

    tz = None
    if ZoneInfo:
        try:
            tz = ZoneInfo(TIMEZONE)
        except Exception:
            tz = None

    now = time.time() if now_ts is None else float(now_ts)
    if tz:
        now_dt = time.localtime(now)
        # ç”¨ tz-aware datetime
        import datetime as _dt
        n = _dt.datetime.fromtimestamp(now, tz)
        year = n.year
        dt = _dt.datetime(year, mon, day, hh, mm, tzinfo=tz)
        if dt.timestamp() <= now:
            dt = _dt.datetime(year + 1, mon, day, hh, mm, tzinfo=tz)
        return dt.timestamp()
    else:
        # é€€åŒ–ï¼šç”¨æœ¬åœ°æ—¶åŒº
        lt = time.localtime(now)
        year = lt.tm_year
        import datetime as _dt
        dt = _dt.datetime(year, mon, day, hh, mm)
        ts = dt.timestamp()
        if ts <= now:
            dt2 = _dt.datetime(year + 1, mon, day, hh, mm)
            ts = dt2.timestamp()
        return ts


def pretty_ts(ts: float) -> str:
    try:
        return time.strftime("%Y-%m-%d %H:%M", time.localtime(ts))
    except Exception:
        return str(ts)


# ========= ä»»åŠ¡æ•°æ®ç»“æ„ =========
@dataclass
class HandinTask:
    task_id: str
    group_id: int
    creator_id: int
    name: str
    created_ts: float
    # å¯é€‰çš„å¤šä¸ªæé†’æ—¶é—´ï¼ˆæ—¶é—´æˆ³åˆ—è¡¨ï¼‰ã€‚æœ€åä¸€ä¸ªæ—¶é—´ä¸€å®šæ˜¯æˆªæ­¢æ—¶é—´ï¼Œæé†’æ—¶é—´å¯ä¸ºç©ºã€‚
    remind_ts_list: List[float] = field(default_factory=list)
    # å·²å‘é€åˆ°ç¬¬å‡ ä¸ªæé†’ï¼ˆä¸‹ä¸€ä¸ªå°†å‘é€çš„æé†’ç´¢å¼•ï¼‰
    remind_sent_idx: int = 0
    deadline_ts: float = 0.0
    deadline_sent: bool = False
    closed: bool = False
    cancelled: bool = False
    cancelled_ts: float = 0.0
    cancelled_by: int = 0

    # ä»»åŠ¡åˆ›å»ºè€…æœ€åä¸€æ¬¡ /handinget çš„æ—¶é—´æˆ³ï¼ˆç”¨äºå½’æ¡£ä¿ç•™ç­–ç•¥ï¼‰
    last_handinget_ts: float = 0.0
    # å½’æ¡£æ˜¯å¦å·²è¢«æ¸…ç†ï¼ˆæ¸…ç†å /handinget ä¸å†å¯ç”¨ï¼Œä½†æ—¥å¿—ä»ä¿ç•™ï¼‰
    purged: bool = False
    purged_ts: float = 0.0

    def is_active(self, now: Optional[float] = None) -> bool:
        now = time.time() if now is None else float(now)
        return (not self.closed) and now < float(self.deadline_ts)


class HandinService:
    """æäº¤ä»»åŠ¡æœåŠ¡ï¼šä»»åŠ¡ç®¡ç† + æäº¤æ–‡ä»¶å½’æ¡£ + æœªäº¤åå•ç»Ÿè®¡ + å®šæ—¶æé†’/æˆªæ­¢æ¨é€ã€‚"""

    def __init__(self, log: Logger):
        self.log = log
        self.db_path = Path(HANDIN_DB_PATH)
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self.inbox_dir = Path(HANDIN_INBOX_DIR)
        self.inbox_dir.mkdir(parents=True, exist_ok=True)

        # æ–°ç‰ˆï¼šæäº¤æ–‡ä»¶ä¸å†æ”¾åœ¨ data/groups/<gid>/handin ä¸‹ï¼Œé¿å…ç¾¤æˆå‘˜é€šè¿‡ /find çœ‹åˆ°ä»–äººæäº¤
        # ç»Ÿä¸€æ”¾åœ¨ data/handin/<gid>/<task>/files/
        self.handin_root = Path(HANDIN_ROOT_DIR)
        self.handin_root.mkdir(parents=True, exist_ok=True)

        # å…¼å®¹è¿ç§»ï¼šæŠŠæ—§ç‰ˆ data/groups/<gid>/handin/<task>/... æ¬åˆ° data/handin/<gid>/<task>/...
        self._migrate_legacy_tree()

        self._tasks: Dict[str, HandinTask] = {}
        self._load()

        # æ¸…ç†èŠ‚æµï¼šé¿å…æ¯ 10 ç§’å…¨ç›˜æ‰«æ
        self._last_cleanup_ts: float = 0.0
        # åå†Œç¼“å­˜ï¼ˆæŒ‰ mtime åˆ·æ–°ï¼‰
        self._roster_cache_mtime: float = -1.0
        self._roster_cache: List[Tuple[str, str]] = []

    def is_task_gettable(self, task: HandinTask) -> bool:
        """ä»»åŠ¡æ˜¯å¦ä»å¯ /handingetï¼šå½’æ¡£æœªè¢«æ¸…ç†ä¸”ç›®å½•ä»åœ¨ã€‚"""
        try:
            if getattr(task, "purged", False):
                return False
            files_dir = self._task_files_dir(task.group_id, task.name)
            return files_dir.exists()
        except Exception:
            return False

    def _purge_task_archive(self, task: HandinTask, now: Optional[float] = None) -> bool:
        """åˆ é™¤æŸä»»åŠ¡çš„å½’æ¡£ç›®å½•ï¼Œå¹¶æ ‡è®°ä¸º purgedã€‚è¿”å›æ˜¯å¦æœ‰å˜æ›´ã€‚"""
        now = time.time() if now is None else float(now)
        try:
            tdir = self._task_dir(task.group_id, task.name)
            if tdir.exists():
                shutil.rmtree(tdir, ignore_errors=True)
        except Exception:
            pass
        changed = False
        if not getattr(task, "purged", False):
            task.purged = True
            task.purged_ts = now
            changed = True
        return changed

    def cleanup_archives_and_inbox(self, now: Optional[float] = None) -> bool:
        """æ¸…ç†ï¼š
        - å½’æ¡£ï¼šåœ¨ä»»åŠ¡åˆ›å»ºè€…æœ€åä¸€æ¬¡ /handinget åä¿ç•™ HANDIN_KEEP_DAYS_AFTER_LAST_GET å¤©
        - inboxï¼šä¸´æ—¶æ”¶ä»¶ç®±å†…æ–‡ä»¶ä¿ç•™ HANDIN_INBOX_KEEP_DAYS å¤©
        è¿”å›ï¼šæ˜¯å¦å‘ç”Ÿäº† DB å˜æ›´ï¼ˆéœ€è¦ä¿å­˜ï¼‰ã€‚
        """
        now = time.time() if now is None else float(now)
        changed = False

        keep_sec = float(HANDIN_KEEP_DAYS_AFTER_LAST_GET) * 86400.0
        # 1) å½’æ¡£æ¸…ç†
        for t in list(self._tasks.values()):
            if getattr(t, "purged", False):
                continue
            last_get = float(getattr(t, "last_handinget_ts", 0.0) or 0.0)
            if last_get <= 0:
                continue
            # ä»åœ¨è¿›è¡Œä¸­çš„ä»»åŠ¡ä¸æ¸…ç†
            if t.is_active(now):
                continue
            if now - last_get >= keep_sec:
                if self._purge_task_archive(t, now=now):
                    changed = True

        # 2) inbox æ¸…ç†ï¼ˆæŒ‰æ–‡ä»¶ mtimeï¼‰
        inbox_keep = float(HANDIN_INBOX_KEEP_DAYS) * 86400.0
        try:
            if self.inbox_dir.exists():
                for p in self.inbox_dir.rglob("*"):
                    if not p.is_file():
                        continue
                    try:
                        if now - float(p.stat().st_mtime) >= inbox_keep:
                            p.unlink(missing_ok=True)
                    except Exception:
                        continue
        except Exception:
            pass

        return changed

    # ----- persistence -----
    def _load(self):
        try:
            if self.db_path.exists():
                obj = json.loads(self.db_path.read_text(encoding="utf-8"))
                if isinstance(obj, dict):
                    for tid, t in obj.items():
                        if not isinstance(t, dict):
                            continue
                        td = dict(t)

                        # å…¼å®¹æ—§å­—æ®µï¼šremind_ts/remind_sent
                        if "remind_ts_list" not in td:
                            r = td.get("remind_ts", None)
                            td["remind_ts_list"] = [float(r)] if r is not None else []
                            if td.get("remind_sent") is True and td["remind_ts_list"]:
                                td["remind_sent_idx"] = len(td["remind_ts_list"])
                            else:
                                td["remind_sent_idx"] = 0
                            td.pop("remind_ts", None)
                            td.pop("remind_sent", None)
                        else:
                            td["remind_ts_list"] = [float(x) for x in (td.get("remind_ts_list") or [])]
                            td.setdefault("remind_sent_idx", 0)

                        td.setdefault("deadline_sent", False)
                        td.setdefault("closed", False)
                        td.setdefault("cancelled", False)
                        td.setdefault("cancelled_ts", 0.0)
                        td.setdefault("cancelled_by", 0)
                        td.setdefault("last_handinget_ts", 0.0)
                        td.setdefault("purged", False)
                        td.setdefault("purged_ts", 0.0)
                        self._tasks[str(tid)] = HandinTask(**td)
        except Exception as e:
            self.log.warning(f"Handin DB load failed: {e}")
            self._tasks = {}

    def _save(self):
        try:
            obj = {tid: asdict(t) for tid, t in self._tasks.items()}
            tmp = self.db_path.with_suffix(self.db_path.suffix + ".tmp")
            tmp.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")
            tmp.replace(self.db_path)
        except Exception as e:
            self.log.warning(f"Handin DB save failed: {e}")

    def _get_roster(self) -> List[Tuple[str, str]]:
        """è¯»å–å¹¶ç¼“å­˜åå†Œï¼ˆæ–‡ä»¶ mtime å˜åŒ–æ—¶è‡ªåŠ¨åˆ·æ–°ï¼‰ã€‚"""
        path = Path(ROSTER_XLSX_PATH)
        if not path.exists():
            self._roster_cache = []
            self._roster_cache_mtime = -1.0
            return []
        try:
            mtime = float(path.stat().st_mtime)
        except Exception:
            mtime = -1.0
        if mtime >= 0 and abs(mtime - float(self._roster_cache_mtime)) < 1e-6:
            return list(self._roster_cache)
        try:
            data = load_roster(path)
        except Exception:
            data = []
        self._roster_cache = list(data or [])
        self._roster_cache_mtime = mtime
        return list(self._roster_cache)

    def _get_roster_names(self) -> List[str]:
        names: List[str] = []
        seen: Set[str] = set()
        for _, nm in self._get_roster():
            name = str(nm or "").strip()
            if (not name) or (name in seen):
                continue
            seen.add(name)
            names.append(name)
        names.sort(key=lambda s: len(s), reverse=True)
        return names

    def find_roster_name_in_filename(self, filename: str, roster_names: Optional[List[str]] = None) -> str:
        """åœ¨æ–‡ä»¶åä¸­æŸ¥æ‰¾æ˜¯å¦åŒ…å«åå†Œä¸­çš„å§“åï¼Œè¿”å›é¦–ä¸ªå‘½ä¸­çš„å§“åã€‚"""
        fn = str(filename or "")
        if not fn:
            return ""
        stem = Path(fn).stem
        compact = re.sub(r"\s+", "", stem)
        names = roster_names if roster_names is not None else self._get_roster_names()
        for nm in names:
            if nm and (nm in stem or nm in compact):
                return nm
        return ""

    # ----- paths -----
    def _task_dir(self, group_id: int, task_name: str) -> Path:
        safe = self._safe_component(task_name)
        return (self.handin_root / str(group_id) / safe)

    def _task_files_dir(self, group_id: int, task_name: str) -> Path:
        d = self._task_dir(group_id, task_name) / "files"
        d.mkdir(parents=True, exist_ok=True)
        return d

    def _legacy_handin_dir(self, group_id: int) -> Path:
        """æ—§ç‰ˆæœ¬ï¼šdata/groups/<gid>/handin"""
        return GROUP_DOCS_DIR / str(group_id) / HANDIN_TASKS_DIRNAME

    @staticmethod
    def _move_or_merge_dir(src: Path, dst: Path, log: Logger) -> None:
        """æŠŠ src ç›®å½•è¿ç§»åˆ° dstã€‚dst ä¸å­˜åœ¨åˆ™ç›´æ¥ moveï¼›å­˜åœ¨åˆ™åˆå¹¶å¹¶å¯¹å†²çªæ–‡ä»¶é‡å‘½åã€‚"""
        src = Path(src)
        dst = Path(dst)
        if not src.exists() or not src.is_dir():
            return
        dst.parent.mkdir(parents=True, exist_ok=True)

        if not dst.exists():
            try:
                shutil.move(str(src), str(dst))
            except Exception as e:
                log.warning(f"[handin] move legacy dir failed: {src} -> {dst}: {e}")
            return

        # dst å·²å­˜åœ¨ï¼šé€’å½’åˆå¹¶
        def merge_dir(a: Path, b: Path):
            b.mkdir(parents=True, exist_ok=True)
            for item in a.iterdir():
                target = b / item.name
                if item.is_dir():
                    merge_dir(item, target)
                    try:
                        item.rmdir()
                    except Exception:
                        pass
                else:
                    if target.exists():
                        stem, suf = target.stem, target.suffix
                        for i in range(1, 999):
                            alt = b / f"{stem}_legacy{i}{suf}"
                            if not alt.exists():
                                target = alt
                                break
                    try:
                        shutil.move(str(item), str(target))
                    except Exception as e:
                        log.warning(f"[handin] move legacy file failed: {item} -> {target}: {e}")

        try:
            merge_dir(src, dst)
            # å°è¯•åˆ é™¤æ®‹ä½™
            try:
                shutil.rmtree(src)
            except Exception:
                pass
        except Exception as e:
            log.warning(f"[handin] merge legacy dir failed: {src} -> {dst}: {e}")

    def _migrate_legacy_tree(self) -> None:
        """å¯åŠ¨æ—¶è¿ç§»æ‰€æœ‰ç¾¤çš„æ—§ handin ç›®å½•åˆ° data/handin ä¸‹ï¼Œå¹¶å°½é‡æ¸…ç†æ—§ç›®å½•ã€‚"""
        try:
            if not GROUP_DOCS_DIR.exists():
                return

            moved_any = False
            for gdir in GROUP_DOCS_DIR.iterdir():
                if not gdir.is_dir():
                    continue
                gid = gdir.name
                legacy = gdir / HANDIN_TASKS_DIRNAME
                if not legacy.exists() or not legacy.is_dir():
                    continue

                dst_gid = self.handin_root / gid
                for task_dir in legacy.iterdir():
                    if not task_dir.is_dir():
                        continue
                    dst_task = dst_gid / task_dir.name
                    self._move_or_merge_dir(task_dir, dst_task, self.log)
                    moved_any = True

                # å°è¯•åˆ é™¤æ—§ handin ç›®å½•ï¼ˆç©ºåˆ™æˆåŠŸï¼‰
                try:
                    legacy.rmdir()
                except Exception:
                    pass

            if moved_any:
                self.log.info("[handin] migrated legacy submissions into data/handin/")
        except Exception as e:
            self.log.warning(f"[handin] legacy migration failed: {e}")

    @staticmethod
    def _safe_component(s: str, max_len: int = 80) -> str:
        s = (s or "").strip()
        s = re.sub(r'[<>:"/\\|?*]', "_", s)
        s = re.sub(r"\s+", " ", s).strip()
        s = s.rstrip(" .")
        if not s:
            s = "_"
        if len(s) > max_len:
            s = s[:max_len].rstrip(" .") or "_"
        return s

    # ----- task ops -----
    def list_active_tasks(self) -> List[HandinTask]:
        now = time.time()
        tasks = [t for t in self._tasks.values() if t.is_active(now)]
        tasks.sort(key=lambda x: x.deadline_ts)
        return tasks

    def list_active_tasks_by_group(self, group_id: int) -> List[HandinTask]:
        return [t for t in self.list_active_tasks() if int(t.group_id) == int(group_id)]


    def list_active_tasks_by_creator(self, creator_id: int) -> List[HandinTask]:
        """åˆ—å‡ºæŸä¸ªå‘èµ·äººåˆ›å»ºçš„æ­£åœ¨è¿›è¡Œä»»åŠ¡ï¼ˆè·¨ç¾¤ï¼‰ã€‚"""
        return [t for t in self.list_active_tasks() if int(t.creator_id) == int(creator_id)]

    # ===== æ–°å¢ï¼šåˆ—å‡ºä»»åŠ¡ï¼ˆåŒ…å«å·²æˆªæ­¢/å·²ç»“æŸ/å·²å–æ¶ˆï¼‰=====
    def list_tasks(self, include_closed: bool = True) -> List[HandinTask]:
        """åˆ—å‡ºä»»åŠ¡ã€‚include_closed=True æ—¶åŒ…å«å·²æˆªæ­¢/å·²ç»“æŸ/å·²å–æ¶ˆçš„ä»»åŠ¡ã€‚"""
        tasks = list(self._tasks.values())
        if not include_closed:
            tasks = [t for t in tasks if not t.closed]
        # è¿‘æœŸä¼˜å…ˆï¼šæŒ‰æˆªæ­¢æ—¶é—´å€’åº
        tasks.sort(key=lambda x: float(x.deadline_ts), reverse=True)
        return tasks

    def list_tasks_by_group(self, group_id: int, include_closed: bool = True) -> List[HandinTask]:
        """åˆ—å‡ºæŸç¾¤çš„ä»»åŠ¡ï¼ˆå«å·²æˆªæ­¢ï¼‰ã€‚"""
        return [t for t in self.list_tasks(include_closed=include_closed) if int(t.group_id) == int(group_id)]

    def list_tasks_by_creator(self, creator_id: int, include_closed: bool = True) -> List[HandinTask]:
        """åˆ—å‡ºæŸä¸ªå‘èµ·äººåˆ›å»ºçš„ä»»åŠ¡ï¼ˆè·¨ç¾¤ï¼Œå«å·²æˆªæ­¢ï¼‰ã€‚"""
        return [t for t in self.list_tasks(include_closed=include_closed) if int(t.creator_id) == int(creator_id)]

    def list_submitted_files(self, task: HandinTask) -> List[Path]:
        """åˆ—å‡ºæŸä»»åŠ¡å·²æäº¤çš„æ–‡ä»¶ï¼ˆæŒ‰ä¿®æ”¹æ—¶é—´å€’åºï¼‰ã€‚"""
        files_dir = self._task_files_dir(task.group_id, task.name)
        if not files_dir.exists():
            return []
        out = [p for p in files_dir.iterdir() if p.is_file()]
        out.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        return out

    def zip_submissions(self, task: HandinTask, out_zip: Path) -> Tuple[bool, str, Optional[Path]]:
        """å°†æŸä»»åŠ¡å·²æäº¤æ–‡ä»¶å…¨éƒ¨æ‰“åŒ…ä¸º zipã€‚"""
        if getattr(task, "purged", False) or (not self._task_files_dir(task.group_id, task.name).exists()):
            return False, "è¯¥ä»»åŠ¡å½’æ¡£å·²è¶…è¿‡ä¿ç•™æœŸï¼ˆæœ€åä¸€æ¬¡ /handinget åå·²æ¸…ç†ï¼‰ï¼Œæ— æ³•å†å¯¼å‡ºã€‚å¦‚éœ€é•¿æœŸä¿ç•™è¯·åŠæ—¶å¤‡ä»½ã€‚", None
        import zipfile
        files = self.list_submitted_files(task)
        out_zip = Path(out_zip)
        out_zip.parent.mkdir(parents=True, exist_ok=True)
        try:
            with zipfile.ZipFile(out_zip, "w", compression=zipfile.ZIP_DEFLATED) as z:
                for p in files:
                    z.write(p, arcname=p.name)
            return True, f"å·²æ‰“åŒ… {len(files)} ä¸ªæ–‡ä»¶ï¼š{out_zip.name}", out_zip
        except Exception as e:
            return False, f"æ‰“åŒ…å¤±è´¥ï¼š{e}", None

    def create_task(self, group_id: int, creator_id: int, name: str, remind_ts_list: Optional[List[float]], deadline_ts: float) -> Tuple[bool, str]:
        """åˆ›å»ºä»»åŠ¡ï¼šæé†’æ—¶é—´å¯ä¸ºç©ºæˆ–å¤šä¸ªï¼›æœ€åä¸€ä¸ªæ—¶é—´ä¸€å®šæ˜¯æˆªæ­¢æ—¶é—´ï¼ˆç”±å‘½ä»¤è§£æä¿è¯ï¼‰ã€‚"""
        name = (name or "").strip()
        if not name or " " in name:
            return False, "ä»»åŠ¡åä¸åˆæ³•ï¼šä¸èƒ½ä¸ºç©ºä¸”ä¸èƒ½åŒ…å«ç©ºæ ¼ã€‚"
        if deadline_ts is None:
            return False, "æ—¶é—´æ ¼å¼ä¸å¯¹ï¼šè¯·ç”¨ æœˆ.æ—¥ æ—¶:åˆ†ï¼Œä¾‹å¦‚ 1.22 18:30ï¼ˆå†’å·ä¸­è‹±æ–‡éƒ½è¡Œï¼‰ã€‚"

        rlist = []
        for x in (remind_ts_list or []):
            try:
                if x is None:
                    continue
                rlist.append(float(x))
            except Exception:
                continue
        # å»é‡å¹¶æ’åº
        if rlist:
            rlist = sorted(set(rlist))

        dts = float(deadline_ts)
        if rlist and rlist[-1] >= dts:
            return False, "æé†’æ—¶é—´å¿…é¡»æ—©äºæˆªæ­¢æ—¶é—´ã€‚"

        # åŒç¾¤åŒåä¸”æœªæˆªæ­¢çš„ä»»åŠ¡ä¸å…è®¸é‡å¤åˆ›å»º
        for t in self._tasks.values():
            if int(t.group_id) == int(group_id) and t.name == name and t.is_active():
                return False, f"ä»»åŠ¡å·²å­˜åœ¨ï¼š{name}ï¼ˆè¯¥ç¾¤å†…åŒåä»»åŠ¡å°šæœªæˆªæ­¢ï¼‰"

        tid = f"{int(group_id)}:{name}:{int(time.time())}"
        task = HandinTask(
            task_id=tid,
            group_id=int(group_id),
            creator_id=int(creator_id),
            name=name,
            created_ts=time.time(),
            remind_ts_list=rlist,
            remind_sent_idx=0,
            deadline_ts=dts,
        )
        self._tasks[tid] = task
        # åˆ›å»ºç›®å½•
        self._task_files_dir(task.group_id, task.name)
        self._save()

        msg_lines = [f"åˆ›å»ºæäº¤ä»»åŠ¡æˆåŠŸï¼š{name}"]
        if task.remind_ts_list:
            for i, ts in enumerate(task.remind_ts_list, 1):
                msg_lines.append(f"æé†’{i}ï¼š{pretty_ts(ts)}")
        else:
            msg_lines.append("æé†’ï¼šæ— ")
        msg_lines.append(f"æˆªæ­¢ï¼š{pretty_ts(task.deadline_ts)}")
        return True, "\n".join(msg_lines)
    
    def cancel_task(self, task_id: str, by_user_id: int) -> Tuple[bool, str]:
        """å–æ¶ˆä»»åŠ¡ï¼šå°†ä»»åŠ¡æ ‡è®°ä¸º closed/cancelledï¼Œåœæ­¢åç»­æé†’ä¸æˆªæ­¢æ¨é€ã€‚"""
        tid = str(task_id)
        t = self._tasks.get(tid)
        if not t:
            return False, "ä»»åŠ¡ä¸å­˜åœ¨ã€‚"
        if t.closed:
            return False, "ä»»åŠ¡å·²ç»“æŸ/å·²å–æ¶ˆã€‚"
        # æ ‡è®°å–æ¶ˆ
        t.closed = True
        t.deadline_sent = True
        t.cancelled = True
        t.cancelled_ts = time.time()
        t.cancelled_by = int(by_user_id)
        self._save()
        return True, f"å·²å–æ¶ˆä»»åŠ¡ã€Œ{t.name}ã€ï¼ˆç¾¤ {t.group_id}ï¼‰ã€‚"

# ----- submissions -----
    def _unique_path(self, dst_dir: Path, filename: str) -> Path:
        filename = self._safe_component(filename, max_len=120)
        p = dst_dir / filename
        if not p.exists():
            return p
        stem = p.stem
        suf = p.suffix
        for i in range(2, 999):
            p2 = dst_dir / f"{stem}_{i}{suf}"
            if not p2.exists():
                return p2
        return dst_dir / f"{stem}_{int(time.time())}{suf}"

    @staticmethod
    def _normalize_download_url(url: str, filename: str) -> str:
        """QQ/FTN ä¸‹è½½é“¾æ¥æœ‰æ—¶ä¼šå¸¦ç©ºçš„ fname= å‚æ•°ï¼›è¡¥å…¨å®ƒå¯æ˜¾è‘—æé«˜å¯ä¸‹è½½æˆåŠŸç‡ã€‚"""
        url = (url or "").strip()
        if not url:
            return ""
        try:
            sp = urllib.parse.urlsplit(url)
            pairs = urllib.parse.parse_qsl(sp.query, keep_blank_values=True)
            has_fname = False
            new_pairs = []
            for k, v in pairs:
                if k == "fname":
                    has_fname = True
                    if not v:
                        new_pairs.append((k, filename))
                    else:
                        new_pairs.append((k, v))
                else:
                    new_pairs.append((k, v))

            # æŸäº›é“¾æ¥æ ¹æœ¬æ²¡æœ‰ fnameï¼Œä½†å®é™…ä¹Ÿéœ€è¦
            if (not has_fname) and ("ftn_handler" in sp.path):
                new_pairs.append(("fname", filename))

            new_q = urllib.parse.urlencode(new_pairs, doseq=True, encoding="utf-8", errors="strict")
            return urllib.parse.urlunsplit((sp.scheme, sp.netloc, sp.path, new_q, sp.fragment))
        except Exception:
            # å…œåº•ï¼šæœ€ç®€å•çš„è¡¥å…¨
            if url.endswith("fname="):
                return url + urllib.parse.quote(filename)
            return url

    @staticmethod
    def _pick_latest_temp_match(temp_dir: Path, *names: str) -> Optional[Path]:
        """åœ¨ NapCat temp é‡ŒæŒ‰æ–‡ä»¶å/å‰ç¼€å…œåº•åŒ¹é…ï¼Œè¿”å›æœ€æ–°æ–‡ä»¶ã€‚"""
        if not temp_dir.exists() or not temp_dir.is_dir():
            return None

        patterns: list[str] = []
        for raw_name in names:
            nm = (raw_name or "").strip()
            if not nm:
                continue
            p = Path(nm)
            patterns.append(p.name)
            if p.suffix:
                patterns.append(f"{p.stem}*{p.suffix}")

        seen = set()
        hits: list[Path] = []
        for pat in patterns:
            if (not pat) or (pat in seen):
                continue
            seen.add(pat)
            try:
                for m in temp_dir.glob(pat):
                    if m.is_file():
                        hits.append(m)
            except Exception:
                continue

        if not hits:
            return None
        hits.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        return hits[0]

    def download_to_inbox(
            self,
            user_id: int,
            fname: str,
            url: str,
            expected_size: Optional[int] = None,
            timeout: float = 180.0,
    ) -> Tuple[bool, str, Optional[Path]]:
        """ä¿å­˜ç§èŠæ–‡ä»¶åˆ° inboxã€‚

        url å¯èƒ½æ˜¯ï¼š
        1) http/https ç›´é“¾ï¼ˆå¯ç›´æ¥ä¸‹è½½ï¼‰
        2) NapCat å®¹å™¨å†…æœ¬åœ°è·¯å¾„ï¼ˆä¾‹å¦‚ /app/.config/QQ/NapCat/temp/xxxï¼‰
        3) file:///... å½¢å¼
        """
        # å…¼å®¹ï¼šfname ä¸ºç©ºæ—¶å…œåº•
        fname = (fname or "file").strip()

        raw = (url or "").strip()
        if not raw:
            return False, "æ–‡ä»¶ç¼ºå°‘ä¸‹è½½é“¾æ¥ï¼ˆurlï¼‰ã€‚", None

        user_dir = self.inbox_dir / str(int(user_id))
        user_dir.mkdir(parents=True, exist_ok=True)

        # ç”Ÿæˆå”¯ä¸€ç›®æ ‡è·¯å¾„
        dst = self._unique_path(user_dir, fname)
        dst_part = dst.with_suffix(dst.suffix + ".part")

        # -------- 1) å¤„ç† file:/// è·¯å¾„ --------
        if raw.startswith("file:///"):
            # file:///C:/xxx æˆ– file:////app/... éƒ½å¯èƒ½å‡ºç°ï¼Œç»Ÿä¸€è½¬æˆæœ¬åœ°è·¯å¾„å­—ç¬¦ä¸²
            u = urlparse(raw)
            raw = unquote(u.path)  # Linux: /app/... ; Windows å¯èƒ½æ˜¯ /C:/...
            if os.name == "nt" and raw.startswith("/") and len(raw) >= 4 and raw[2] == ":":
                raw = raw[1:]  # /C:/xx -> C:/xx

        # -------- 2) å¤„ç†å®¹å™¨å†…æœ¬åœ°ç¼“å­˜è·¯å¾„ï¼ˆä»¥ / å¼€å¤´ï¼‰--------
        if raw.startswith("/"):
            try:
                temp_dir = Path(NAPCAT_TEMP_HOST_DIR)
                cdir = str(NAPCAT_TEMP_CONTAINER_DIR).rstrip("/")
                # è‹¥è·¯å¾„åœ¨ NapCat temp ä¸‹ï¼ŒæŒ‰æ˜ å°„å…³ç³»æ‰¾åˆ°å®¿ä¸»æœºå¯¹åº”æ–‡ä»¶
                if raw.startswith(cdir + "/") or raw == cdir:
                    rel = raw[len(cdir):].lstrip("/")
                    src = temp_dir / rel
                else:
                    # å…œåº•ï¼šæŒ‰ basename åœ¨ temp ç›®å½•æ‰¾
                    src = temp_dir / Path(raw).name

                # NapCat äº‹ä»¶æœ‰æ—¶æ—©äºç¼“å­˜è½ç›˜ï¼Œç»™å‡ ç§’ç­‰å¾…å¹¶åšä¸€æ¬¡æ¨¡ç³ŠåŒ¹é…å…œåº•
                deadline = time.time() + 8.0
                while not (src.exists() and src.is_file()):
                    alt = self._pick_latest_temp_match(temp_dir, Path(raw).name, fname)
                    if alt is not None:
                        src = alt
                        break
                    if time.time() >= deadline:
                        break
                    time.sleep(0.4)

                if not src.exists() or not src.is_file():
                    if not temp_dir.exists():
                        return False, f"ä¸‹è½½æ–‡ä»¶å¤±è´¥ï¼šNapCat æœ¬åœ°ç¼“å­˜ç›®å½•ä¸å­˜åœ¨ï¼š{temp_dir}ï¼ˆè¯·æ£€æŸ¥ NAPCAT_TEMP_HOST_DIRï¼‰", None
                    return False, f"ä¸‹è½½æ–‡ä»¶å¤±è´¥ï¼šNapCat æœ¬åœ°ç¼“å­˜æ–‡ä»¶ä¸å­˜åœ¨ï¼š{src}", None

                # å¤§æ–‡ä»¶å¯èƒ½æ­£åœ¨è½ç›˜ï¼šç¨ç­‰ï¼Œç­‰ size æœ‰æ˜æ˜¾å¢é•¿/è¾¾åˆ°ä¸€å®šæ¯”ä¾‹
                for _ in range(12):
                    if expected_size:
                        try:
                            exp = int(expected_size)
                            if exp > 0 and src.stat().st_size < max(32, exp // 10):
                                time.sleep(0.5)
                                continue
                        except Exception:
                            pass
                    break

                shutil.copy2(src, dst)
                size = dst.stat().st_size
                return True, f"å·²æ”¶åˆ°æ–‡ä»¶ï¼š{dst.name}ï¼ˆ{size} bytesï¼Œæœ¬åœ°ç¼“å­˜æ‹·è´ï¼‰", dst
            except Exception as e:
                return False, f"ä¸‹è½½æ–‡ä»¶å¤±è´¥ï¼šæœ¬åœ°ç¼“å­˜æ‹·è´å¼‚å¸¸ï¼š{e}", None

        # -------- 3) http/https ä¸‹è½½ --------
        if raw.startswith("http://") or raw.startswith("https://"):
            try:
                req = urllib.request.Request(
                    raw,
                    headers={
                        "User-Agent": "Mozilla/5.0",
                    },
                )

                # ç”¨ .part å†™å…¥ï¼Œé¿å…åŠæˆªæ–‡ä»¶è¢«å½“æˆæˆåŠŸ
                if dst_part.exists():
                    try:
                        dst_part.unlink()
                    except Exception:
                        pass

                with urllib.request.urlopen(req, timeout=float(timeout)) as resp, open(dst_part, "wb") as f:
                    downloaded = 0
                    chunk = 1024 * 1024  # 1MB
                    while True:
                        b = resp.read(chunk)
                        if not b:
                            break
                        f.write(b)
                        downloaded += len(b)

                        # å¦‚æœ expected_size å·²çŸ¥ï¼Œä¸”ä¸€ç›´å‡ ä¹ä¸å¢é•¿ï¼Œä¹Ÿå¯ä»¥ç»™æ—¶é—´è®©ç½‘ç»œç¼“å†²
                        # è¿™é‡Œä¸ä¸»åŠ¨ä¸­æ–­ï¼Œåªåšç¨³å¦¥å†™å…¥

                # ä¸‹è½½å®Œæˆåæ”¹å
                os.replace(dst_part, dst)

                size = dst.stat().st_size
                # è‹¥ expected_size å­˜åœ¨ä¸”å·®è·ç¦»è°±ï¼Œæç¤ºå¯èƒ½ä¸å®Œæ•´
                if expected_size:
                    try:
                        exp = int(expected_size)
                        if exp > 0 and size < exp * 0.5:
                            return False, f"ä¸‹è½½ç–‘ä¼¼ä¸å®Œæ•´ï¼šæœŸæœ›çº¦ {exp} bytesï¼Œå®é™… {size} bytesï¼ˆå¯èƒ½é“¾æ¥å¤±æ•ˆ/è¢«æ‹¦æˆªï¼‰", None
                    except Exception:
                        pass

                return True, f"å·²æ”¶åˆ°æ–‡ä»¶ï¼š{dst.name}ï¼ˆ{size} bytesï¼Œç½‘ç»œä¸‹è½½ï¼‰", dst
            except Exception as e:
                # æ¸…ç† part
                try:
                    if dst_part.exists():
                        dst_part.unlink()
                except Exception:
                    pass
                return False, f"ä¸‹è½½æ–‡ä»¶å¤±è´¥ï¼šç½‘ç»œä¸‹è½½å¼‚å¸¸ï¼š{e}", None

        return False, f"ä¸æ”¯æŒçš„ä¸‹è½½æ¥æºï¼š{raw}", None


    def move_inbox_to_task(self, inbox_path: Path, task: HandinTask, overwrite: bool = False) -> Tuple[bool, str, Optional[Path], str]:
        """å°† inbox ä¸´æ—¶æ–‡ä»¶ç§»åŠ¨åˆ°ä»»åŠ¡ files ç›®å½•ã€‚

        - è‹¥ç›®æ ‡å­˜åœ¨åŒåæ–‡ä»¶ä¸” overwrite=Falseï¼šè¿”å› code='EXISTS' å¹¶ä¸ç§»åŠ¨æ–‡ä»¶
        - overwrite=Trueï¼šè¦†ç›–ç›®æ ‡æ–‡ä»¶
        """
        if not inbox_path or not Path(inbox_path).exists():
            return False, "ä¸´æ—¶æ–‡ä»¶ä¸å­˜åœ¨ï¼ˆå¯èƒ½å·²è¿‡æœŸ/è¢«æ¸…ç†ï¼‰ã€‚", None, "MISSING"

        dst_dir = self._task_files_dir(task.group_id, task.name)
        dst_dir.mkdir(parents=True, exist_ok=True)
        dst = dst_dir / Path(inbox_path).name

        if dst.exists() and not overwrite:
            return False, f"ä»»åŠ¡ã€Œ{task.name}ã€ä¸­å·²å­˜åœ¨åŒåæ–‡ä»¶ï¼š{dst.name}", dst, "EXISTS"

        try:
            if dst.exists() and overwrite:
                dst.unlink()
            Path(inbox_path).replace(dst)
            return True, f"å·²å½’æ¡£åˆ°ä»»åŠ¡ã€Œ{task.name}ã€ï¼š{dst.name}", dst, "OK"
        except Exception as e:
            return False, f"å½’æ¡£å¤±è´¥ï¼š{e}", None, "ERR"
    # ----- roster compare -----
    def compute_missing(self, task: HandinTask) -> Tuple[bool, str, List[Tuple[str, str]], Dict]:
        """è¿”å› (ok, msg, missing_list, stats)."""
        roster = self._get_roster()
        if not roster:
            return False, f"è¯»å–ç­çº§åå†Œå¤±è´¥ï¼š{ROSTER_XLSX_PATH}ï¼ˆæ–‡ä»¶ä¸å­˜åœ¨æˆ–æ ¼å¼ä¸å¯¹ï¼‰", [], {}

        submitted_ids: Set[str] = set()
        submitted_names: Set[str] = set()
        submitted_file_names: List[str] = []
        unknown_name_files: List[str] = []
        matched_name_files = 0

        roster_name_set = {str(nm or "").strip() for _, nm in roster if str(nm or "").strip()}
        roster_names = sorted(roster_name_set, key=lambda s: len(s), reverse=True)

        for p in self.list_submitted_files(task):
            # ç»Ÿè®¡æ‰€æœ‰å·²æäº¤æ–‡ä»¶ï¼›ä»…è·³è¿‡éšè—æ–‡ä»¶ä¸ä¸´æ—¶åˆ†ç‰‡
            if p.name.startswith("."):
                continue
            if p.suffix.lower() == ".part":
                continue
            submitted_file_names.append(p.name)

            sid = extract_student_id(p.name)
            if sid:
                submitted_ids.add(sid)

            nm = self.find_roster_name_in_filename(p.name, roster_names=roster_names)
            if not nm:
                # å…¼å®¹æ—§è§„åˆ™ï¼šå…ˆæŠ½å–å§“åï¼Œå†æ£€æŸ¥æ˜¯å¦ç¡®å®åœ¨åå†Œä¸­
                nm_guess = extract_name_from_filename(p.name)
                if nm_guess and (nm_guess in roster_name_set):
                    nm = nm_guess

            if nm:
                submitted_names.add(nm)
                matched_name_files += 1
            else:
                unknown_name_files.append(p.name)

        missing = []
        handed = 0
        for sid, nm in roster:
            if (sid and sid in submitted_ids) or (nm and nm in submitted_names):
                handed += 1
            else:
                missing.append((sid, nm))

        stats = {
            "roster_total": len(roster),
            "handed_in": handed,
            "missing": len(missing),
            "submitted_ids": len(submitted_ids),
            "submitted_names": len(submitted_names),
            "submitted_files_total": len(submitted_file_names),
            "recognized_name_files": matched_name_files,
            "recognized_name_ratio": (float(matched_name_files) / float(len(submitted_file_names))) if submitted_file_names else 1.0,
            "unknown_name_files": unknown_name_files,
            "submitted_file_names": submitted_file_names,
        }
        ratio = float(stats.get("recognized_name_ratio", 1.0))
        total_files = int(stats.get("submitted_files_total", 0))
        stats["use_submitted_list"] = bool(total_files > 0 and (matched_name_files <= 0 or ratio < 0.2))
        return True, "ok", missing, stats

    def format_missing_message(self, task: HandinTask, missing: List[Tuple[str, str]], stats: Dict, title: str) -> str:
        lines: List[str] = []
        lines.append(f"{title}\nä»»åŠ¡ï¼š{task.name}\nç¾¤ï¼š{task.group_id}\næˆªæ­¢ï¼š{pretty_ts(task.deadline_ts)}")
        lines.append(f"å·²äº¤/æ€»äººæ•°ï¼š{stats.get('handed_in',0)}/{stats.get('roster_total',0)}ï¼›æœªäº¤ï¼š{stats.get('missing',0)}")
        total_files = int(stats.get("submitted_files_total", 0) or 0)
        matched_name_files = int(stats.get("recognized_name_files", 0) or 0)
        ratio = float(stats.get("recognized_name_ratio", 0.0) or 0.0)
        if total_files > 0:
            lines.append(f"å§“åè¯†åˆ«æ–‡ä»¶å æ¯”ï¼š{matched_name_files}/{total_files}ï¼ˆ{ratio * 100:.1f}%ï¼‰")

        submitted_file_names = list(stats.get("submitted_file_names") or [])
        unknown_name_files = list(stats.get("unknown_name_files") or [])
        use_submitted_list = bool(stats.get("use_submitted_list"))

        # æ–‡ä»¶åè¯†åˆ«ç‡å¤ªä½æ—¶ï¼Œæœªäº¤åå•å‡†ç¡®æ€§ä¸è¶³ï¼Œæ”¹å‘å·²äº¤æ–‡ä»¶åˆ—è¡¨
        if use_submitted_list:
            lines.append("âš ï¸ å§“åè¯†åˆ«ç‡è¿‡ä½ï¼ˆ<20%ï¼‰æˆ–æœªè¯†åˆ«åˆ°åå†Œå§“åï¼Œæ”¹ä¸ºå‘é€å·²æäº¤æ–‡ä»¶åˆ—è¡¨ã€‚")
            if not submitted_file_names:
                lines.append("å½“å‰æ²¡æœ‰å·²æäº¤æ–‡ä»¶ã€‚")
                return "\n".join(lines)
            lines.append("å·²æäº¤æ–‡ä»¶åˆ—è¡¨ï¼š")
            limit_files = 120
            for i, fn in enumerate(submitted_file_names[:limit_files], 1):
                lines.append(f"{i}. {fn}")
            if len(submitted_file_names) > limit_files:
                lines.append(f"...ï¼ˆå…± {len(submitted_file_names)} ä¸ªï¼Œå·²æˆªæ–­æ˜¾ç¤ºå‰ {limit_files} ä¸ªï¼‰")
            return "\n".join(lines)

        if not missing:
            lines.append("âœ… å…¨éƒ¨å·²æäº¤ã€‚")
        else:
            lines.append("æœªäº¤åå•ï¼š")
            limit = 120
            for i, (sid, nm) in enumerate(missing[:limit], 1):
                if nm:
                    lines.append(f"{i}. {nm}")
                else:
                    lines.append(f"{i}. ï¼ˆæœªçŸ¥ï¼‰")
            if len(missing) > limit:
                lines.append(f"...ï¼ˆå…± {len(missing)} äººï¼Œå·²æˆªæ–­æ˜¾ç¤ºå‰ {limit} äººï¼‰")

        # é¢å¤–åˆ—å‡ºâ€œå·²æäº¤ä½†æœªè¯†åˆ«å‡ºå§“åä¿¡æ¯â€çš„æ–‡ä»¶å
        if unknown_name_files:
            lines.append("")
            lines.append("æœªè¯†åˆ«åˆ°å§“åä¿¡æ¯çš„å·²æäº¤æ–‡ä»¶ï¼š")
            limit_unknown = 80
            for i, fn in enumerate(unknown_name_files[:limit_unknown], 1):
                lines.append(f"{i}. {fn}")
            if len(unknown_name_files) > limit_unknown:
                lines.append(f"...ï¼ˆå…± {len(unknown_name_files)} ä¸ªï¼Œå·²æˆªæ–­æ˜¾ç¤ºå‰ {limit_unknown} ä¸ªï¼‰")
        return "\n".join(lines)

    # ----- scheduler -----
    async def scheduler_loop(self, api):
        """å®šæ—¶æ£€æŸ¥æé†’/æˆªæ­¢ã€‚"""
        while True:
            try:
                await asyncio.sleep(10)
                now = time.time()
                changed = False

                # å‘¨æœŸæ€§æ¸…ç†ï¼ˆå½’æ¡£ + inboxï¼‰ã€‚é»˜è®¤æ¯ 1 å°æ—¶æœ€å¤šè·‘ä¸€æ¬¡ã€‚
                try:
                    if now - float(self._last_cleanup_ts or 0.0) >= 3600.0:
                        if self.cleanup_archives_and_inbox(now=now):
                            changed = True
                        self._last_cleanup_ts = now
                except Exception:
                    pass
                for t in list(self._tasks.values()):
                    if t.closed:
                        continue

                    # remind (0~N æ¬¡)
                    while t.remind_sent_idx < len(t.remind_ts_list) and now >= float(t.remind_ts_list[t.remind_sent_idx]):
                        idx = int(t.remind_sent_idx)
                        title = "ğŸ“Œ ä½œä¸šæäº¤æé†’"
                        if len(t.remind_ts_list) > 1:
                            title = f"ğŸ“Œ ä½œä¸šæäº¤æé†’ï¼ˆç¬¬ {idx+1}/{len(t.remind_ts_list)} æ¬¡ï¼‰"
                        ok, msg, missing, stats = self.compute_missing(t)
                        if ok:
                            text = self.format_missing_message(t, missing, stats, title)
                        else:
                            text = title + "\n" + msg
                        await api.send_private_msg(t.creator_id, text)
                        t.remind_sent_idx = idx + 1
                        changed = True

                    # deadline
                    if (not t.deadline_sent) and now >= float(t.deadline_ts):
                        ok, msg, missing, stats = self.compute_missing(t)
                        if ok:
                            text = self.format_missing_message(t, missing, stats, "â° ä½œä¸šæˆªæ­¢æé†’ï¼ˆå·²åˆ°æˆªæ­¢æ—¶é—´ï¼‰")
                        else:
                            text = "â° ä½œä¸šæˆªæ­¢æé†’\n" + msg
                        await api.send_private_msg(t.creator_id, text)
                        t.deadline_sent = True
                        t.closed = True
                        changed = True

                if changed:
                    self._save()
            except Exception as e:
                self.log.warning(f"handin scheduler error: {e}")
